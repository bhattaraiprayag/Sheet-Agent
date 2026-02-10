"""
Deterministic A/R Aging Report Generator.

This module contains the pure Python logic for generating Accounts Receivable
aging reports. It replaces the non-deterministic LLM-based approach with
testable, reproducible calculations.
"""

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def generate_ar_aging_report(
    excel_path: str,
    reporting_date: str,
    column_map: dict[str, str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Generates an Accounts Receivable (A/R) Aging Report from an Excel sheet.

    This function implements deterministic business logic for identifying
    cumulative rows, invoice rows, and credit rows, then calculates maturity
    clusters and aggregations.

    Args:
        excel_path: The path to the input Excel file.
        reporting_date: The date to use for maturity calculations (format: YYYY-MM-DD).
        column_map: Dictionary mapping semantic keys to actual column names.
                   Expected keys: 'amount_local_currency', 'due_date', 'assignment',
                   'posting_date', 'document_type'.

    Returns:
        A tuple containing two DataFrames:
        - detailed_data: The original data with added analysis columns
        - summary_report: The final aging report with aggregated metrics

    Raises:
        FileNotFoundError: If the Excel file doesn't exist.
        KeyError: If required columns are missing from column_map.
    """
    try:
        df = pd.read_excel(excel_path)
        logger.info(f"Successfully loaded Excel file: {excel_path}")
    except FileNotFoundError:
        logger.error(f"File not found: {excel_path}")
        raise
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        raise

    # Extract column names from the semantic mapping
    amt_col = column_map["amount_local_currency"]
    date_col = column_map["due_date"]
    zuordnung_col = column_map["assignment"]
    posting_date_col = column_map["posting_date"]
    doc_type_col = column_map["document_type"]

    logger.info(f"Using column mappings: {column_map}")

    # --- Find the cutoff point ---
    # Identify the index of the first row containing 'Hauptbuchkonto' in the assignment column.
    # All calculations will be stopped for rows at and after this point.
    hauptbuch_rows = df[df[zuordnung_col].astype(str).str.contains("Hauptbuchkonto", na=False)]
    stop_index = hauptbuch_rows.index.min() if not hauptbuch_rows.empty else len(df)
    active_mask = df.index < stop_index
    logger.info(f"Cutoff point (Hauptbuchkonto) found at index: {stop_index}")

    # --- 1. Identify Cumulative Rows ---
    s, cumul = 0, []

    for index, r in df.iterrows():
        # If the row is at or after the stop_index, append NaN and skip calculation.
        if not active_mask[index]:
            cumul.append(np.nan)
            continue

        if pd.isna(r[amt_col]):
            cumul.append(False)
            continue

        matches = pd.isna(r[date_col]) and abs(r[amt_col] - s) < 0.01 and s != 0
        is_summary = False
        if zuordnung_col in df.columns and isinstance(r[zuordnung_col], str):
            is_summary = any(
                x in r[zuordnung_col] for x in ["Debitor", "Hauptbuch", "Buchungskreis"]
            )
        c = matches and is_summary
        cumul.append(c)
        s = 0 if c else s + r[amt_col]

    # Using nullable boolean type to handle True/False/NA
    df["Cumulative"] = pd.Series(cumul, dtype="boolean")
    logger.info(f"Identified {df['Cumulative'].sum()} cumulative rows")

    # --- 2. Feature Engineering ---
    # Identify invoice rows: non-cumulative rows with positive amounts
    invoice_condition = (df[posting_date_col].notna()) & (df[amt_col] >= 0)
    df["Invoice"] = np.where(active_mask, invoice_condition, np.nan)
    df["Invoice"] = df["Invoice"].astype("boolean")

    # Move Cumulative column next to Invoice for better visibility
    cumulative_col_data = df.pop("Cumulative")
    df.insert(df.columns.get_loc("Invoice"), "Cumulative", cumulative_col_data)

    # Identify credit rows: non-cumulative rows with negative amounts
    credit_condition = (df[doc_type_col].notna()) & (df[amt_col] <= 0)
    df["Credit"] = np.where(active_mask, credit_condition, np.nan)
    df["Credit"] = df["Credit"].astype("boolean")

    logger.info(
        f"Identified {df['Invoice'].sum()} invoice rows and {df['Credit'].sum()} credit rows"
    )

    # Calculate Due Date and Maturity
    df["Due Date"] = pd.to_datetime(df[date_col], errors="coerce")
    day_diff = (df["Due Date"] - pd.to_datetime(reporting_date)).dt.days

    # Calculate maturity values first, then mask them to leave trailing rows empty.
    is_valid_transaction = (df["Invoice"].fillna(False) | df["Credit"].fillna(False)) & (
        df["Due Date"].notna()
    )
    maturity_values = np.where(is_valid_transaction, day_diff, -6)
    df["Maturity"] = np.where(active_mask, maturity_values, np.nan)

    # Create maturity clusters
    conditions = [
        df["Maturity"] < -60,
        (df["Maturity"] >= -60) & (df["Maturity"] < -30),
        (df["Maturity"] >= -30) & (df["Maturity"] < 0),
    ]
    choices = [">60 days", "31-60 days", "1-30 days"]

    # Calculate cluster values, then mask them to leave trailing rows empty.
    cluster_values = np.select(conditions, choices, default="Not mature")
    cluster_values_filtered = np.where(
        df["Invoice"].fillna(False) | df["Credit"].fillna(False), cluster_values, None
    )
    df["Cluster"] = np.where(active_mask, cluster_values_filtered, None)

    # --- 3. Aggregation and Reporting ---
    cluster_categories = ["Not mature", "1-30 days", "31-60 days", ">60 days"]

    # Filter using .fillna(False) to handle the new <NA> values correctly.
    invoice_summary = (
        df[df["Invoice"].fillna(False)]
        .groupby("Cluster")[amt_col]
        .sum()
        .reindex(cluster_categories)
        .fillna(0)
    )
    credit_summary = (
        df[df["Credit"].fillna(False)]
        .groupby("Cluster")[amt_col]
        .sum()
        .reindex(cluster_categories)
        .fillna(0)
    )

    total_invoice = invoice_summary.sum()
    total_credit = credit_summary.sum()

    invoice_percentages = (invoice_summary / total_invoice) if total_invoice != 0 else 0
    credit_percentages = (credit_summary / total_credit) if total_credit != 0 else 0

    logger.info(f"Total invoice amount: {total_invoice}, Total credit amount: {total_credit}")

    summary_report = pd.DataFrame(
        {
            "Sum of Invoice Amounts": [total_invoice] + [np.nan] * (len(invoice_summary) - 1),
            "Sum of Credit Amounts": [total_credit] + [np.nan] * (len(credit_summary) - 1),
            "(Invoice) Maturity Cluster": invoice_summary.index,
            "Total Amount": invoice_summary.values,
            "Percentage": invoice_percentages.values,
            "(Credit) Maturity Cluster": credit_summary.index,
            "Total Amount_crd": credit_summary.values,
            "Percentage_crd": credit_percentages.values,
        }
    )

    # Standardize column names (handle duplicate "Total Amount" and "Percentage")
    summary_report.columns = [
        "Sum of Invoice Amounts",
        "Sum of Credit Amounts",
        "(Invoice) Maturity Cluster",
        "Total Amount",
        "Percentage",
        "(Credit) Maturity Cluster",
        "Total Amount",
        "Percentage",
    ]

    # Add row number references
    row_num_df = pd.DataFrame(
        {
            "Cumulative Row Numbers": pd.Series(
                [i + 2 for i in df.index[df["Cumulative"].fillna(False)]]
            ),
            "Invoice Row Numbers": pd.Series(
                [i + 2 for i in df.index[df["Invoice"].fillna(False)]]
            ),
            "Credit Row Numbers": pd.Series([i + 2 for i in df.index[df["Credit"].fillna(False)]]),
        }
    )
    summary_report = pd.concat([summary_report, row_num_df], axis=1)

    logger.info("Successfully generated A/R aging report")
    return df, summary_report


def _format_worksheet(worksheet, dataframe, format_map=None):
    """
    Helper function to auto-fit and format columns in an openpyxl worksheet.

    Args:
        worksheet: The openpyxl worksheet to format.
        dataframe: The source DataFrame to determine column widths.
        format_map: Optional dictionary mapping column indices to number formats.
    """
    if format_map is None:
        format_map = {}
    center_align = Alignment(horizontal="center", vertical="center")

    for i, col_name in enumerate(dataframe.columns, 1):
        header_len = len(str(col_name))
        data_series = dataframe.iloc[:, i - 1].dropna()
        data_len = data_series.astype(str).map(len).max() if not data_series.empty else 0
        width = max(header_len, data_len) + 2
        worksheet.column_dimensions[get_column_letter(i)].width = width

        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=i)
            cell.alignment = center_align
            if i in format_map:
                cell.number_format = format_map[i]


def _create_processed_sheet(
    output_path: str, detailed_data: pd.DataFrame, hide_sheet: bool = False
) -> str:
    """
    Creates the 'Processed_' sheet by copying the original sheet to preserve
    formatting, adds new calculated columns, and optionally hides the sheet.

    Args:
        output_path: Path to the output Excel file.
        detailed_data: DataFrame containing the detailed analysis data.
        hide_sheet: Whether to hide the processed sheet in the final workbook.

    Returns:
        The name of the created processed sheet.
    """
    wb = load_workbook(output_path)
    original_sheet_name = wb.sheetnames[0]
    detailed_sheet_name = f"Processed_{original_sheet_name}"

    if detailed_sheet_name in wb.sheetnames:
        wb.remove(wb[detailed_sheet_name])

    source_ws = wb[original_sheet_name]
    processed_ws = wb.copy_worksheet(source_ws)
    processed_ws.title = detailed_sheet_name

    new_columns_to_add = ["Cumulative", "Invoice", "Credit", "Due Date", "Maturity", "Cluster"]
    start_col_idx = processed_ws.max_column + 1
    center_align = Alignment(horizontal="center", vertical="center")

    for i, col_name in enumerate(new_columns_to_add):
        current_col_idx = start_col_idx + i
        processed_ws.cell(row=1, column=current_col_idx, value=col_name)
        for row_idx, value in enumerate(detailed_data[col_name], start=2):
            cell_value = (
                bool(value) if isinstance(value, np.bool_) else (None if pd.isna(value) else value)
            )
            cell = processed_ws.cell(row=row_idx, column=current_col_idx, value=cell_value)
            cell.alignment = center_align
            if isinstance(cell_value, pd.Timestamp):
                cell.number_format = "YYYY-MM-DD"
    if hide_sheet:
        processed_ws.sheet_state = "hidden"

    wb.save(output_path)
    logger.info(f"Created {'hidden ' if hide_sheet else ''}processed sheet: {detailed_sheet_name}")
    return detailed_sheet_name


def _create_analysis_sheet(output_path: str, final_report: pd.DataFrame, currency_symbol: str):
    """
    Adds the 'Analysis' sheet to the workbook and moves it to the second position.

    Args:
        output_path: Path to the output Excel file.
        final_report: DataFrame containing the final aging report.
        currency_symbol: The currency symbol to use in formatting (e.g., '€', '$').
    """
    with pd.ExcelWriter(
        output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        final_report.to_excel(writer, sheet_name="Analysis", index=False)

        # Apply currency-specific number formatting
        summary_formats = {
            1: f"{currency_symbol} #,##0.00",
            2: f"{currency_symbol} #,##0.00",
            4: f"{currency_symbol} #,##0.00",
            7: f"{currency_symbol} #,##0.00",
            5: "0.00%",
            8: "0.00%",
        }
        _format_worksheet(writer.sheets["Analysis"], final_report, summary_formats)

        wb = writer.book
        wb.move_sheet("Analysis", offset=-1)

    logger.info(f"Created Analysis sheet with currency symbol: {currency_symbol}")


def create_ar_report(
    input_path: Path,
    output_path: Path,
    reporting_date: str,
    column_map: dict[str, str],
    currency_symbol: str,
    hide_processed_sheet: bool = True,
) -> None:
    """
    Main entry point for creating the A/R aging report.

    This function coordinates the entire report generation process:
    1. Generates the detailed and summary data
    2. Creates the processed sheet (with optional hiding)
    3. Creates the Analysis sheet with proper formatting

    Args:
        input_path: Path to the input Excel file.
        output_path: Path where the output Excel file will be saved.
        reporting_date: The date to use for maturity calculations (format: YYYY-MM-DD).
        column_map: Dictionary mapping semantic keys to actual column names.
        currency_symbol: The currency symbol to use in formatting.
        hide_processed_sheet: Whether to hide the processed sheet in the final workbook.

    Raises:
        Exception: If report generation fails at any step.
    """
    logger.info(f"Starting A/R report generation: {input_path} -> {output_path}")

    try:
        # Generate the detailed and summary data
        detailed_data, final_report = generate_ar_aging_report(
            excel_path=str(output_path), reporting_date=reporting_date, column_map=column_map
        )

        if final_report.empty:
            raise ValueError("Report generation failed: No data was produced.")

        # Create the Processed sheet (and hide if configured)
        _create_processed_sheet(str(output_path), detailed_data, hide_sheet=hide_processed_sheet)

        # Create the Analysis sheet and reorder it
        _create_analysis_sheet(str(output_path), final_report, currency_symbol)

        logger.info(f"Report successfully generated: {output_path}")

    except Exception as e:
        logger.error(f"Error during report generation: {e}")
        raise
